from dataclasses import dataclass
from datetime import datetime, timedelta
from decimal import Decimal
import csv
import os
import re
import xml.etree.ElementTree as ET
from typing import List, Dict, Any, Optional

try:
    import pandas as pd  # type: ignore
except Exception:
    pd = None


DATE_FORMAT = "%Y-%m-%d"


def _to_date_str(value: str) -> Optional[str]:
    if not value:
        return None
    text = str(value).strip()
    # try ISO-like first
    m = re.search(r"(\d{4})[-/](\d{2})[-/](\d{2})", text)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = re.search(r"(\d{2})[-/](\d{2})[-/](\d{4})", text)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    # try datetime parser fallback
    try:
        dt = datetime.fromisoformat(text)
        return dt.date().isoformat()
    except Exception:
        pass
    try:
        dt = datetime.strptime(text, "%d/%m/%Y")
        return dt.date().isoformat()
    except Exception:
        pass
    return None


def _clean_number(value: Any) -> Optional[float]:
    if value is None:
        return None
    s = str(value).strip()
    if s == "":
        return None
    s = s.replace("$", "").replace("\u00A0", "").replace(",", "")
    m = re.search(r"-?\d+\.?\d*", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _normalize_row(row: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    # Try to find keys for date, description/concept, reference, amount, debit/credit
    keys = {k.lower().strip(): v for k, v in row.items()}
    date_keys = ["fecha", "date", "fechaoperacion", "fechamovimiento", "fecha_operacion"]
    desc_keys = ["descripcion", "concepto", "detalle", "description", "memo"]
    ref_keys = ["referencia", "ref", "folio", "autorizacion", "autorización", "id"]
    amount_keys = ["monto", "importe", "amount", "valor", "cantidad"]
    debit_keys = ["cargo", "debito", "débito", "retiro"]
    credit_keys = ["abono", "deposito", "depósito", "credito", "crédito", "ingreso"]

    def pick(cands):
        for c in cands:
            if c in keys and keys[c] not in (None, ""):
                return keys[c]
        return None

    fecha = _to_date_str(str(pick(date_keys)) if pick(date_keys) is not None else "")
    descripcion = pick(desc_keys) or pick(["concepto"])
    referencia = pick(ref_keys)
    monto = None
    tipo = None

    # If separate debit/credit columns
    credit = pick(credit_keys)
    debit = pick(debit_keys)
    if credit not in (None, ""):
        c = _clean_number(credit)
        if c is not None and c > 0:
            tipo = "abono"
            monto = c
    if debit not in (None, "") and monto is None:
        d = _clean_number(debit)
        if d is not None and d > 0:
            tipo = "cargo"
            monto = d
    # fallback to single amount column
    if monto is None:
        amt = pick(amount_keys)
        if amt is not None:
            val = _clean_number(amt)
            if val is not None:
                if val < 0:
                    tipo = "cargo"
                    monto = abs(val)
                else:
                    tipo = "abono"
                    monto = val

    if fecha is None or monto is None:
        return None

    return {
        "Fecha": fecha,
        "Concepto": str(descripcion or "").strip(),
        "Referencia": str(referencia or "").strip() if referencia is not None else None,
        "Monto": round(float(monto), 2),
        "Tipo": tipo,
    }


def _read_csv(path: str) -> List[Dict[str, Any]]:
    results = []
    with open(path, newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            results.append(row)
    return results


def _read_xlsx(path: str) -> List[Dict[str, Any]]:
    # prefer pandas if available
    if pd is not None:
        df = pd.read_excel(path, dtype=str)
        return df.fillna("").to_dict(orient="records")
    # fallback to openpyxl
    try:
        from openpyxl import load_workbook
    except Exception:
        raise RuntimeError("openpyxl is required to read xlsx files")
    wb = load_workbook(filename=path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    results = []
    for r in rows[1:]:
        d = {headers[i]: r[i] for i in range(len(headers))}
        results.append(d)
    return results


def _read_xml(path: str) -> List[Dict[str, Any]]:
    with open(path, "rb") as fh:
        xml_bytes = fh.read()
    root = ET.fromstring(xml_bytes)
    records = []
    # walk tree and build dicts from element tags/attrib/text
    for node in root.iter():
        data = {}
        # attributes
        for k, v in node.attrib.items():
            data[k.lower()] = v
        # children text
        for child in list(node):
            text = (child.text or "").strip()
            if text:
                data[child.tag.lower()] = text
            for k, v in child.attrib.items():
                data[k.lower()] = v
        # own text
        if (node.text or "").strip():
            data[node.tag.lower()] = (node.text or "").strip()
        if data:
            records.append(data)
    return records


def _read_pdf_text_records(path: str) -> List[Dict[str, Any]]:
    def _extract_text_pages(path: str) -> list[str | None]:
        try:
            from PyPDF2 import PdfReader
            reader = PdfReader(path)
            return [page.extract_text() for page in reader.pages]
        except Exception:
            pass

        try:
            from pdfminer.high_level import extract_text
            text = extract_text(path)
            return [text] if text is not None else []
        except Exception as exc:
            raise RuntimeError(
                "La lectura de PDF requiere pdfplumber, PyPDF2 o pdfminer.six. Instala una de estas dependencias."
            ) from exc

    records: List[Dict[str, Any]] = []
    text_pages = _extract_text_pages(path)
    for text in text_pages:
        if not text:
            continue
        lines = [line.strip() for line in text.splitlines() if line.strip()]
        if not lines:
            continue
        header_start = None
        for i, line in enumerate(lines):
            if any(word in line.lower() for word in ("fecha", "date")):
                header_start = i
                break
        if header_start is None:
            continue

        headers = [lines[header_start].strip().lower()]
        data_start = header_start + 1
        while data_start < len(lines):
            candidate = lines[data_start]
            if re.search(r"\d{4}[-/]\d{2}[-/]\d{2}", candidate) or re.search(r"\d{2}[-/]\d{2}[-/]\d{4}", candidate):
                break
            headers.append(candidate.strip().lower())
            data_start += 1

        rest_lines = lines[data_start:]
        if not headers or not rest_lines:
            continue

        for line in rest_lines:
            parts = re.split(r"\s{2,}", line)
            if len(parts) >= 2:
                record = {
                    headers[j]: parts[j].strip() if j < len(headers) else ""
                    for j in range(len(headers))
                }
                if any(record.values()):
                    records.append(record)

        if not records and len(rest_lines) >= len(headers):
            for idx in range(0, len(rest_lines), len(headers)):
                row_lines = rest_lines[idx : idx + len(headers)]
                if len(row_lines) < len(headers):
                    break
                record = {headers[j]: row_lines[j].strip() for j in range(len(headers))}
                if any(record.values()):
                    records.append(record)
    return records


def _read_pdf(path: str) -> List[Dict[str, Any]]:
    # optional dependency: pdfplumber or tabula
    try:
        import pdfplumber  # type: ignore
    except Exception:
        return _read_pdf_text_records(path)
    records = []
    try:
        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                try:
                    tables = page.extract_tables()
                except Exception:
                    tables = None
                if not tables:
                    continue
                for table in tables:
                    # first row header
                    if not table or len(table) < 2:
                        continue
                    headers = [str(h).strip().lower() for h in table[0]]
                    for row in table[1:]:
                        if not any(row):
                            continue
                        record = {
                            headers[i]: str(row[i]).strip() if i < len(row) and row[i] is not None else ""
                            for i in range(len(headers))
                        }
                        records.append(record)
    except Exception:
        return _read_pdf_text_records(path)

    if records:
        return records
    return _read_pdf_text_records(path)


def cargar_estados_cuenta(ruta_archivo: str) -> List[Dict[str, Any]]:
    """
    Cargar estados de cuenta desde CSV, XLSX, XML o PDF.
    Devuelve una lista de filas normalizadas con columnas: Fecha (YYYY-MM-DD), Concepto, Referencia, Monto (float), Tipo (cargo/abono)
    """
    if not os.path.exists(ruta_archivo):
        raise FileNotFoundError(ruta_archivo)
    ext = os.path.splitext(ruta_archivo)[1].lower()
    raw_records: List[Dict[str, Any]] = []
    if ext in {".csv", ".txt"}:
        raw_records = _read_csv(ruta_archivo)
    elif ext in {".xlsx", ".xls"}:
        raw_records = _read_xlsx(ruta_archivo)
    elif ext in {".xml"}:
        raw_records = _read_xml(ruta_archivo)
    elif ext in {".pdf"}:
        raw_records = _read_pdf(ruta_archivo)
    else:
        raise RuntimeError(f"Formato de archivo no soportado: {ext}")

    normalized = []
    for r in raw_records:
        try:
            nr = _normalize_row(r)
            if nr:
                normalized.append(nr)
        except Exception:
            # skip rows that fail normalization
            continue
    return normalized


def conciliar_movimientos(estados_cuenta: List[Dict[str, Any]], xml_polizas: List[bytes]) -> Dict[str, Any]:
    """
    Conciliar movimientos bancarios (lista de dicts) contra pólizas en XML (lista de bytes/strings).
    Coincidencias por fecha (±2 días), monto exacto y referencia parcial.
    Retorna dict con keys: conciliados (list), pendientes_banco (list), pendientes_polizas (list)
    """
    # parse polizas xmls into simplified records
    polizas = []
    for b in xml_polizas:
        try:
            root = ET.fromstring(b)
        except Exception:
            continue
        # try to find nodes that contain amount, date, reference; fallback to root attribs
        for node in root.iter():
            data = {}
            for k, v in node.attrib.items():
                data[k.lower()] = v
            for child in list(node):
                if child.text and child.text.strip():
                    data[child.tag.lower()] = child.text.strip()
            if data:
                fecha = None
                for fk in ("fecha", "date", "fecha_poliza", "fechaemision"):
                    if fk in data:
                        fecha = _to_date_str(data[fk])
                        break
                referencia = None
                for rk in ("referencia", "folio", "uuid", "uuidtimbre", "serie"):
                    if rk in data:
                        referencia = data[rk]
                        break
                monto = None
                for ak in ("monto", "importe", "total", "valor", "amount"):
                    if ak in data:
                        monto = _clean_number(data[ak])
                        if monto is not None:
                            break
                if fecha and monto is not None:
                    polizas.append({"fecha": fecha, "referencia": referencia or "", "monto": round(float(monto), 2), "raw": data})
    pendientes_polizas = polizas.copy()

    conciliados = []
    banca_pendientes = []

    def within_days(d1: str, d2: str, days: int = 2) -> bool:
        try:
            a = datetime.fromisoformat(d1).date()
            b = datetime.fromisoformat(d2).date()
            return abs((a - b).days) <= days
        except Exception:
            return False

    for mov in estados_cuenta:
        matched = None
        for p in pendientes_polizas:
            if round(float(mov.get("Monto", 0)), 2) == round(float(p["monto"]), 2):
                if within_days(mov["Fecha"], p["fecha"], 2):
                    # reference partial match if both present
                    ref_a = (mov.get("Referencia") or "").lower()
                    ref_b = (p.get("referencia") or "").lower()
                    if not ref_a or not ref_b or ref_a in ref_b or ref_b in ref_a:
                        matched = p
                        break
        if matched:
            conciliados.append({"movimiento": mov, "poliza": matched})
            pendientes_polizas.remove(matched)
        else:
            banca_pendientes.append(mov)

    return {
        "conciliados": conciliados,
        "pendientes_banco": banca_pendientes,
        "pendientes_polizas": pendientes_polizas,
    }


def validar_polizas(polizas: List[Dict[str, Any]]) -> Dict[str, List[str]]:
    """
    Verificar y validar pólizas contables en base a movimientos conciliados (simplified).
    Revisa campos obligatorios y reglas simples fiscales/contables.
    Retorna dict: {"ok": [...], "errores": [...], "advertencias": [...]}
    """
    ok = []
    errores = []
    adv = []
    for p in polizas:
        # expect fecha, monto, referencia present
        missing = []
        if not p.get("fecha"):
            missing.append("fecha")
        if p.get("monto") in (None, "", 0):
            missing.append("monto")
        if missing:
            errores.append(f"Poliza missing: {', '.join(missing)} -> {p}")
            continue
        # monto must be positive
        try:
            if float(p.get("monto", 0)) <= 0:
                errores.append(f"Monto no positivo en poliza: {p}")
                continue
        except Exception:
            errores.append(f"Monto inválido en poliza: {p}")
            continue
        # Fecha reasonable
        try:
            _ = datetime.fromisoformat(p["fecha"]).date()
        except Exception:
            adv.append(f"Fecha inválida en poliza: {p}")
        ok.append(p)
    return {"ok": ok, "errores": errores, "advertencias": adv}


def validador_datos_contables(datos: List[Dict[str, Any]]) -> List[str]:
    """
    Revisa integridad de datos contables. Devuelve lista de advertencias.
    """
    warnings = []
    for d in datos:
        if d.get("Monto") is None:
            warnings.append(f"Fila sin monto: {d}")
        else:
            try:
                if float(d.get("Monto")) == 0:
                    warnings.append(f"Monto cero en fila: {d}")
            except Exception:
                warnings.append(f"Monto inválido en fila: {d}")
        if not d.get("Fecha"):
            warnings.append(f"Fila sin fecha: {d}")
        else:
            # date not in future
            try:
                dt = datetime.fromisoformat(d.get("Fecha")).date()
                if dt > datetime.now().date() + timedelta(days=1):
                    warnings.append(f"Fecha futura en fila: {d}")
            except Exception:
                warnings.append(f"Fecha inválida en fila: {d}")
    return warnings


def cargar_constancia_fiscal(ruta_archivo: str) -> Dict[str, Any]:
    """
    Cargar constancia de situación fiscal (XML o PDF). Extrae regimen fiscal y RFC/Nombre si es posible.
    """
    if not os.path.exists(ruta_archivo):
        raise FileNotFoundError(ruta_archivo)
    ext = os.path.splitext(ruta_archivo)[1].lower()
    if ext == ".xml":
        with open(ruta_archivo, "rb") as fh:
            b = fh.read()
        try:
            root = ET.fromstring(b)
        except Exception:
            raise RuntimeError("XML inválido")
        data = {}
        for node in root.iter():
            for k, v in node.attrib.items():
                kl = k.lower()
                if kl in ("rfc", "rfcactivo", "regimen", "regimenfiscal", "regimenfiscalid"):
                    data[kl] = v
            if node.text and node.text.strip():
                # attempt to capture name
                if node.tag.lower() in ("nombre", "nombrefiscal", "name"):
                    data["nombre"] = node.text.strip()
        return {"raw": data}
    else:
        # PDFs not implemented fully; return placeholder or raise informative error
        raise RuntimeError("Carga de PDF para constancias no implementada sin dependencias adicionales")
